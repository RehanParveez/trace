from __future__ import annotations
from decimal import Decimal
from io import BytesIO
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from app.modules.drawings_boq.models import BOQItem, BOQVersion
from app.modules.drawings_boq.words import rupees_in_words
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

_SECTION_TITLES = {
  "MATERIAL": "A. Materials",
  "LABOUR": "B. Labour",
  "CUSTOM": "C. Additional Works",
}

def _group_items(items: list[BOQItem]) -> dict[str, list[BOQItem]]:
  groups: dict[str, list[BOQItem]] = {"MATERIAL": [], "LABOUR": [], "CUSTOM": []}
  for item in items:
    groups[item.item_type.value].append(item)
  return groups

def _section_total(items: list[BOQItem]) -> Decimal:
  return sum(
    (item.quantity * item.unit_rate for item in items if item.unit_rate is not None),
    Decimal("0"),
  )

def build_boq_pdf(
  boq_version: BOQVersion,
  items: list[BOQItem],
  company_name: str,
) -> bytes:
  buffer = BytesIO()
  doc = SimpleDocTemplate(
    buffer, pagesize=A4,
    topMargin=18 * mm, bottomMargin=18 * mm,
    leftMargin=16 * mm, rightMargin=16 * mm,
  )
  styles = getSampleStyleSheet()
  meta = boq_version.export_meta or {}
  company = meta.get("company_name") or company_name

  story = [
    Paragraph(company, ParagraphStyle("Co", parent=styles["Heading1"], fontSize=16)),
    Paragraph("BILL OF QUANTITIES", styles["Heading2"]),
  ]

  if any(item.status.value == "DRAFT" for item in items):
    story.append(Paragraph(
      "DRAFT — contains unapproved line items, figures may change.",
      ParagraphStyle("Warn", parent=styles["Normal"], textColor=colors.red),
    ))

  project_rows = [
    ["Client", meta.get("client_name", "-"), "Project Title", meta.get("project_title", "-")],
    ["Location", meta.get("location", "-"), "Plot Size", meta.get("plot_size", "-")],
    ["Covered Area (Sft)", str(boq_version.covered_area_sqft or "-"), "Storeys", meta.get("storeys", "-")],
    ["Date", boq_version.created_at.strftime("%d %B %Y"), "", ""],
  ]
  project_table = Table(project_rows, colWidths=[35 * mm, 55 * mm, 35 * mm, 55 * mm])
  project_table.setStyle(TableStyle([
    ("FONTSIZE", (0, 0), (-1, -1), 9),
    ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
    ("FONTNAME", (2, 0), (2, -1), "Helvetica-Bold"),
    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
  ]))
  story += [Spacer(1, 6), project_table, Spacer(1, 10)]

  groups = _group_items(items)
  grand_total = Decimal("0")

  for key in ("MATERIAL", "LABOUR", "CUSTOM"):
    section_items = groups[key]
    if not section_items:
      continue
    story.append(Paragraph(_SECTION_TITLES[key], styles["Heading3"]))
    rows = [["#", "Description", "Unit", "Qty", "Rate (Rs)", "Amount (Rs)"]]
    for idx, item in enumerate(section_items, start=1):
      amount = item.quantity * item.unit_rate if item.unit_rate is not None else None
      label = item.material_name + (" (unapproved)" if item.status.value == "DRAFT" else "")
      rows.append([
        str(idx), label, item.unit, f"{item.quantity:,.2f}",
        f"{item.unit_rate:,.2f}" if item.unit_rate is not None else "—",
        f"{amount:,.2f}" if amount is not None else "—",
      ])
    section_total = _section_total(section_items)
    grand_total += section_total
    rows.append(["", "", "", "", "Sub-total", f"{section_total:,.2f}"])

    table = Table(rows, colWidths=[8 * mm, 65 * mm, 15 * mm, 22 * mm, 25 * mm, 30 * mm])
    table.setStyle(TableStyle([
      ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e293b")),
      ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
      ("FONTSIZE", (0, 0), (-1, -1), 8),
      ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
      ("GRID", (0, 0), (-1, -2), 0.25, colors.HexColor("#cbd5e1")),
      ("ALIGN", (3, 0), (-1, -1), "RIGHT"),
    ]))
    story += [table, Spacer(1, 8)]

  cost_per_sqft = (
    grand_total / boq_version.covered_area_sqft
    if boq_version.covered_area_sqft and boq_version.covered_area_sqft > 0
    else None
  )

  story.append(Paragraph(f"<b>GRAND TOTAL: Rs {grand_total:,.2f}</b>", styles["Heading2"]))
  story.append(Paragraph(f"In words: {rupees_in_words(grand_total)}", styles["Normal"]))
  if cost_per_sqft is not None:
    story.append(Paragraph(f"Cost per Sft: Rs {cost_per_sqft:,.2f}", styles["Normal"]))

  story.append(Spacer(1, 20))
  sign_table = Table(
    [["Prepared By", meta.get("prepared_by", "____________________"),
      "Checked By", meta.get("checked_by", "____________________")]],
    colWidths=[25 * mm, 60 * mm, 25 * mm, 60 * mm],
  )
  sign_table.setStyle(TableStyle([("FONTSIZE", (0, 0), (-1, -1), 9)]))
  story.append(sign_table)

  doc.build(story)
  return buffer.getvalue()

def build_boq_xlsx(
  boq_version: BOQVersion,
  items: list[BOQItem],
  company_name: str,
) -> bytes:

  meta = boq_version.export_meta or {}
  company = meta.get("company_name") or company_name

  wb = Workbook()
  ws = wb.active
  ws.title = "Bill of Quantities"

  header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
  header_font = Font(color="FFFFFF", bold=True)
  bold = Font(bold=True)

  ws.append([company])
  ws["A1"].font = Font(bold=True, size=14)
  ws.append(["BILL OF QUANTITIES"])
  ws.append([])
  ws.append(["Client", meta.get("client_name", "-"), "Project Title", meta.get("project_title", "-")])
  ws.append(["Location", meta.get("location", "-"), "Plot Size", meta.get("plot_size", "-")])
  ws.append(["Covered Area (Sft)", str(boq_version.covered_area_sqft or "-"), "Storeys", meta.get("storeys", "-")])
  ws.append([])

  groups = _group_items(items)
  grand_total = Decimal("0")

  for key in ("MATERIAL", "LABOUR", "CUSTOM"):
    section_items = groups[key]
    if not section_items:
      continue
    ws.append([_SECTION_TITLES[key]])
    ws.cell(row=ws.max_row, column=1).font = bold
    header_row = ws.max_row + 1
    ws.append(["#", "Description", "Unit", "Qty", "Rate (Rs)", "Amount (Rs)"])
    for cell in ws[header_row]:
      cell.fill = header_fill
      cell.font = header_font
    for idx, item in enumerate(section_items, start=1):
      amount = item.quantity * item.unit_rate if item.unit_rate is not None else None
      label = item.material_name + (" (unapproved)" if item.status.value == "DRAFT" else "")
      ws.append([
        idx, label, item.unit, float(item.quantity),
        float(item.unit_rate) if item.unit_rate is not None else None,
        float(amount) if amount is not None else None,
      ])
    section_total = _section_total(section_items)
    grand_total += section_total
    ws.append(["", "", "", "", "Sub-total", float(section_total)])
    ws.cell(row=ws.max_row, column=5).font = bold
    ws.cell(row=ws.max_row, column=6).font = bold
    ws.append([])

  cost_per_sqft = (
    grand_total / boq_version.covered_area_sqft
    if boq_version.covered_area_sqft and boq_version.covered_area_sqft > 0
    else None
  )

  ws.append(["GRAND TOTAL", float(grand_total)])
  ws.cell(row=ws.max_row, column=1).font = bold
  ws.append(["In words", rupees_in_words(grand_total)])
  if cost_per_sqft is not None:
    ws.append(["Cost per Sft", float(cost_per_sqft)])
  ws.append([])
  ws.append(["Prepared By", meta.get("prepared_by", ""), "Checked By", meta.get("checked_by", "")])

  for col, width in zip("ABCDEF", (20, 40, 12, 14, 14, 16)):
    ws.column_dimensions[col].width = width

  buffer = BytesIO()
  wb.save(buffer)
  return buffer.getvalue()